#!/usr/bin/env python3
# ============================================================
# golden_compare.py  —  Phase 1 ゴールデン第一判定（セル内容ベース）
#
# 目的:
#   2つの xlsx（ゴールデン=移行前 / 現在=移行後）を比較し、
#   「セル内容ベースの第一判定」で差分ゼロかどうかを判定する。
#
# 判定対象（第一判定・必須）:
#   - 対象シート集合（名前・順序）
#   - 各セルの値
#   - 各セルの数式（openpyxl の data_only=False で取得される '=' 始まりの文字列）
#
# 注意:
#   - xlsx バイナリ一致は判定に使わない（タイムスタンプ/ZIP順/内部IDで揺れるため）。
#   - 翻訳候補1/2/3・既定選択・UiRow など、UIシートの内容もセル値として比較対象に含まれる
#     （確認用UIワークブックはこれらをセルに書き出しているため）。
#
# 使い方:
#   python3 golden_compare.py <golden.xlsx> <current.xlsx>
#   終了コード 0 = 差分ゼロ（合格） / 1 = 差分あり（不合格） / 2 = 実行エラー
# ============================================================

import sys
from openpyxl import load_workbook


def extract(path):
    """xlsx からシート順とセル(値/数式)を抽出して比較可能な構造にする。"""
    wb = load_workbook(path, data_only=False, read_only=True)
    sheets = []
    cells = {}  # (sheet, coordinate) -> normalized value
    for ws in wb.worksheets:  # worksheets はブック内の順序を保持
        sheets.append(ws.title)
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                v = c.value
                # 数値の 1 と "1" を区別したいので型タグを付ける
                cells[(ws.title, c.coordinate)] = f"{type(v).__name__}:{v}"
    wb.close()
    return sheets, cells


def main():
    if len(sys.argv) != 3:
        print("usage: python3 golden_compare.py <golden.xlsx> <current.xlsx>")
        return 2

    golden_path, current_path = sys.argv[1], sys.argv[2]

    try:
        g_sheets, g_cells = extract(golden_path)
        c_sheets, c_cells = extract(current_path)
    except Exception as e:  # noqa: BLE001
        print(f"[ERROR] failed to read: {e}")
        return 2

    diffs = []

    # 1) シート集合・順序
    if g_sheets != c_sheets:
        diffs.append(f"[SHEETS] golden={g_sheets} current={c_sheets}")

    # 2) セル差分
    all_keys = set(g_cells) | set(c_cells)
    cell_diff_count = 0
    for key in sorted(all_keys):
        gv = g_cells.get(key)
        cv = c_cells.get(key)
        if gv != cv:
            cell_diff_count += 1
            if cell_diff_count <= 50:  # 最初の50件だけ詳細表示
                sheet, coord = key
                diffs.append(f"[CELL] {sheet}!{coord} golden={gv!r} current={cv!r}")

    print("=" * 60)
    print("Phase 1 ゴールデン第一判定（セル内容ベース）")
    print(f"  golden : {golden_path}  (sheets={len(g_sheets)}, cells={len(g_cells)})")
    print(f"  current: {current_path}  (sheets={len(c_sheets)}, cells={len(c_cells)})")
    print(f"  cell diffs: {cell_diff_count}")
    print("=" * 60)

    if diffs:
        for d in diffs:
            print(d)
        if cell_diff_count > 50:
            print(f"... and {cell_diff_count - 50} more cell diffs")
        print("RESULT = FAIL (差分あり)")
        return 1

    print("RESULT = PASS (差分ゼロ)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
